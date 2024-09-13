import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value=data
    workbook.save(path)

def get_data_from_excel(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2, total_rows+1):
        row_list = []
        for c in range(1, total_columns+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list

def get_credentials_by_name(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2, total_rows + 1):
        row_dict = {}
        for c in range(1, total_columns + 1):
            if c == 1:
                # Assuming the first column is 'Name'
                row_dict['name'] = sheet.cell(row=r, column=c).value
            elif c == 2:
                # Assuming the second column is 'Username'
                row_dict['username'] = sheet.cell(row=r, column=c).value
            elif c == 3:
                # Assuming the third column is 'Password'
                row_dict['password'] = sheet.cell(row=r, column=c).value
            else:
                # You can add more columns here if needed
                pass
        final_list.append(row_dict)

    return final_list


def get_credentials_by_approver(path, sheet_name, name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row

    for r in range(2, total_rows + 1):
        if sheet.cell(row=r, column=1).value == name:
            username = sheet.cell(row=r, column=2).value
            password = sheet.cell(row=r, column=3).value
            return username, password

    # If the name is not found, return None for both username and password
    return None, None

def get_credentials(path, sheet_name, partial_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row

    for r in range(2, total_rows + 1):
        name = sheet.cell(row=r, column=1).value
        if partial_name.lower() in name.lower():
            username = sheet.cell(row=r, column=2).value
            password = sheet.cell(row=r, column=3).value
            return username, password

    # If no partial match is found, return None for both username and password
    return None, None

def read_LeaveConfig_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
        data.append({
            "max_leaves": row[0],
            "apply_before_days": row[1]
        })
    return data

def get_User_excel(file_path, sheet_name):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if df.empty:
            print(f"Warning: Excel sheet '{sheet_name}' is empty.")
            return None

        # Ensure columns 'Name', 'Userid', 'Password' exist
        if 'Name' not in df.columns or 'Username' not in df.columns or 'Password' not in df.columns:
            print(f"Error: Required columns ('Name', 'Username', 'Password') not found in sheet '{sheet_name}'.")
            return None

        data = df[['Name', 'Username', 'Password']].values.tolist()
        return data
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

